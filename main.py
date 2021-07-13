
import os
from datetime import date
from tabulate import tabulate as tab

import pandas as pd

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from mpl_toolkits.basemap import Basemap

import numpy as np

from PIL import Image, ImageOps

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

today = date.today()

shp_simple_countries = r'C:/Users/Georges/PycharmProjects/data/simple_countries/simple_countries'
shp_simple_areas = r'C:/Users/Georges/PycharmProjects/data/simple_areas/simple_areas'
inputCountryConversion = r'C:/Users/Georges/PycharmProjects/data/countries_conversion.xlsx'

workDirectory = r'C:/Users/Georges/Downloads/'
outputExcelFile = workDirectory+str(today)+' Stats AMS Users XXX.xlsx'

# For now from an Excel import, later we will use the API
inputExcelFile = workDirectory+'User_export_'+str(today)+'.xlsx'
df = pd.read_excel(inputExcelFile, sheet_name='Export', engine='openpyxl',
                   usecols=['ID', 'Email', 'Not blocked', 'Created at', 'Account activation date', 'Live Location:Country',
                            'Industries:Industries', 'Groups Member:Group Member',
                            '_8f70fe1e_Occupation', '_ed5be3a0_How_did_you_hear_about_us_', 'Most Recent Membership Subscription:Type name', 'Most Recent Membership Subscription:Expires at',
                            '_a7634d0d_Professional_Designation__Degree__', '_83fb023d_How_many_years_have_you_been_in_practice_'
                            ])


# FIX COUNTRIES
df['Live Location:Country'] = df['Live Location:Country'].replace(['Hong Kong SAR'], 'Hong Kong')


# COUNTRY-AREA CONVERSION IMPORT
df_CountryConversion = pd.read_excel(inputCountryConversion, sheet_name='countries', engine='openpyxl',
                   usecols=['COUNTRY_HB', 'continent_stat'])


# COUNT ACTIVATION
activeUsers = df[df['Not blocked'] == True].count()['Account activation date']
nonActiveUsers = df[(df['Not blocked'] == True) & (df['Account activation date'].isna())].count()['Not blocked']
blockedConfirmed = df[df['Not blocked'] == False].count()['Account activation date']
blockedUnconfirmed = df[(df['Not blocked'] == False) & (df['Account activation date'].isna())].count()['Not blocked']
allUsers = df['ID'].count()

activationLabel = []
myCounts = []
activationLabel.extend(('Confirmed & not blocked', 'Unconfirmed & not blocked', 'Blocked but confirmed', 'Blocked and unconfirmed', 'Total'))
myCounts.extend((activeUsers, nonActiveUsers, blockedConfirmed, blockedUnconfirmed, allUsers))

ActivationDict = list(zip(activationLabel, myCounts))
df_ActivationCount = pd.DataFrame(ActivationDict, columns =['Users', 'Total'])

df_ActivationCount['%'] = (df_ActivationCount['Total'] / allUsers) * 100
df_ActivationCount['%'] = df_ActivationCount['%'].round(decimals=2)


# COUNT REGISTRATIONS BY DATE (FIELD Created at)
df['Created'] = pd.to_datetime(df['Created at'])
df['Created'] = df['Created'].dt.to_period("M")
df_TempCreated = pd.DataFrame(df['Created'])

df_Created_count = pd.DataFrame(df_TempCreated.groupby(['Created'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Created'], ascending=True).reset_index()

df_Created_count['Created'] = df_Created_count['Created'].dt.strftime('%b %Y')

ind_drop = df_Created_count[df_Created_count['Created'].apply(lambda x: x.startswith('Feb 2020'))].index
df_Created_count = df_Created_count.drop(ind_drop)


# COUNT COUNTRY
df_Country_count = pd.DataFrame(df.groupby(['Live Location:Country'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Country_count = df_Country_count.fillna('Unknow')

df_Country_count['Percent'] = (df_Country_count['Total'] / df_Country_count['Total'].sum()) * 100
df_Country_count['Percent'] = df_Country_count['Percent'].round(decimals=1)


# COUNT AREAS
# JOIN LEFT WITH COUNTRY CONVERSION
df_UsersAreas = pd.merge(df, df_CountryConversion, left_on='Live Location:Country', right_on='COUNTRY_HB', how='left')\
    [['Email', 'continent_stat']]

df_AreasCount = pd.DataFrame(df_UsersAreas.groupby(['continent_stat'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_AreasCount = df_AreasCount.fillna('Unknow')

df_AreasCount['Percent'] = (df_AreasCount['Total'] / df_AreasCount['Total'].sum()) * 100
df_AreasCount['Percent'] = df_AreasCount['Percent'].round(decimals=1)


# COUNT CATEGORIES (CUSTOM FIELD _8f70fe1e_Occupation)
df['Categories'] = df['_8f70fe1e_Occupation'].str.split(': ').str[0]
df_Categories_count = pd.DataFrame(df.groupby(['Categories'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Categories_count = df_Categories_count.fillna('Unknow')

df_Categories_count['Percent'] = (df_Categories_count['Total'] / df_Categories_count['Total'].sum()) * 100
df_Categories_count['Percent'] = df_Categories_count['Percent'].round(decimals=1)


# COUNT SPECIALTIES (CUSTOM FIELD _8f70fe1e_Occupation)
df['Specialties'] = df['_8f70fe1e_Occupation'].str.split(': ').str[1]
df_Specialties_count = pd.DataFrame(df.groupby(['Specialties'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Specialties_count = df_Specialties_count.fillna('Unknow')

df_Specialties_count['Percent'] = (df_Specialties_count['Total'] / df_Specialties_count['Total'].sum()) * 100
df_Specialties_count['Percent'] = df_Specialties_count['Percent'].round(decimals=1)


# COUNT SPECIALTIES PER COUNTRY
df_SpecialtiesPerCountry_count = pd.DataFrame(df.groupby(['Live Location:Country', 'Specialties'], dropna=False)\
    .size(), columns=['Total']).sort_values(['Live Location:Country', 'Total'], ascending=[True, False]).reset_index()
df_SpecialtiesPerCountry_count = df_SpecialtiesPerCountry_count.fillna('Unknow')

df_SpecialtiesPerCountry_count['Percent'] = (df_SpecialtiesPerCountry_count['Total'] / df_SpecialtiesPerCountry_count['Total'].sum()) * 100
df_SpecialtiesPerCountry_count['Percent'] = df_SpecialtiesPerCountry_count['Percent'].round(decimals=2)


# COUNT EXPERTISE & INTERESTS (CUSTOM FIELD Industries:Industries)
df_tempIndustries = pd.DataFrame(pd.melt(df['Industries:Industries'].str.split(',', expand=True))['value'])

# FIX ERRORS
ind_drop = df_tempIndustries[df_tempIndustries['value'].apply(lambda x: x == ('W'))].index
df_tempIndustries = df_tempIndustries.drop(ind_drop)

df_Industries_count = pd.DataFrame(df_tempIndustries.groupby(['value'], dropna=False).size(), columns=['Total'])\
    .reset_index()
df_Industries_count = df_Industries_count.fillna('AZERTY')

df_Industries_count['Percent'] = (df_Industries_count['Total'] / df.shape[0]) * 100
df_Industries_count['Percent'] = df_Industries_count['Percent'].round(decimals=2)

# EMPTY VALUES
industriesEmpty = df['Industries:Industries'].isna().sum()
industriesEmptyPercent = round((industriesEmpty / df.shape[0]) * 100, 2)

# REPLACE EMPTY VALUES AND SORT
df_Industries_count.loc[(df_Industries_count['value'] == 'AZERTY')] = [['Unknow', industriesEmpty, industriesEmptyPercent]]
df_Industries_count = df_Industries_count.sort_values(['Total'], ascending=False)


# COUNT GROUPS (FIELD Groups Member:Group Member)

# The line above to convert in str when all values are number (without "," when no multi-group users)
#df['Groups Member:Group Member'] = df['Groups Member:Group Member'].astype('Int64').astype('str')

df_tempGroups = pd.DataFrame(pd.melt(df['Groups Member:Group Member'].str.split(',', expand=True))['value'])
df_Groups_count = pd.DataFrame(df_tempGroups.groupby(['value'], dropna=False).size(), columns=['Total'])\
    .reset_index()
df_Groups_count = df_Groups_count.fillna('AZERTY')

df_Groups_count['Percent'] = (df_Groups_count['Total'] / df.shape[0]) * 100
df_Groups_count['Percent'] = df_Groups_count['Percent'].round(decimals=2)

# EMPTY VALUES
groupsEmpty = df['Groups Member:Group Member'].isna().sum()
groupsEmptyPercent = round((groupsEmpty / df.shape[0]) * 100, 2)

# REPLACE EMPTY VALUES AND SORT
df_Groups_count.loc[(df_Groups_count['value'] == 'AZERTY')] = [['None', groupsEmpty, groupsEmptyPercent]]
df_Groups_count = df_Groups_count.sort_values(['Total'], ascending=False)
df_Groups_count['value'] = df_Groups_count['value'].replace(['17794'], 'AMS North America')
df_Groups_count['value'] = df_Groups_count['value'].replace(['19659'], 'AMS Asia')
df_Groups_count['value'] = df_Groups_count['value'].replace(['25496'], 'AMS Latin America Chapter')
df_Groups_count['value'] = df_Groups_count['value'].replace(['27725'], 'AMS Eastern Europe (CIS)')
df_Groups_count['value'] = df_Groups_count['value'].replace(['22580'], 'Euro Aesthetics')
df_Groups_count['value'] = df_Groups_count['value'].replace(['24594'], 'Aptos')
df_Groups_count['value'] = df_Groups_count['value'].replace(['25340'], 'Lutronic')
df_Groups_count['value'] = df_Groups_count['value'].replace(['27724'], 'FillMed Laboratoires')
df_Groups_count['value'] = df_Groups_count['value'].replace(['28210'], 'QuantifiCare')
df_Groups_count['value'] = df_Groups_count['value'].replace(['19859'], 'Medicinae Doctor')
df_Groups_count['value'] = df_Groups_count['value'].replace(['24262'], 'Other')
df_Groups_count['value'] = df_Groups_count['value'].replace(['19858'], 'Industries')


# COUNT EMAIL DOMAINS
df['Domain'] = df['Email'].str.split('@').str[1]
df_Email_DNS_count = pd.DataFrame(df.groupby(['Domain'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Email_DNS_count = df_Email_DNS_count.fillna('Unknow')

df_Email_DNS_count['Percent'] = (df_Email_DNS_count['Total'] / df_Email_DNS_count['Total'].sum()) * 100
df_Email_DNS_count['Percent'] = df_Email_DNS_count['Percent'].round(decimals=1)


# COUNT HOW DID YOU HEAR ABOUT US (CUSTOM FIELD How_did_you_hear_about_us_)
df_HowDidYouHearAboutUs_count = pd.DataFrame(df.groupby(['_ed5be3a0_How_did_you_hear_about_us_'], dropna=True).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()

df_HowDidYouHearAboutUs_count['Percent'] = (df_HowDidYouHearAboutUs_count['Total'] / df_HowDidYouHearAboutUs_count['Total'].sum()) * 100
df_HowDidYouHearAboutUs_count['Percent'] = df_HowDidYouHearAboutUs_count['Percent'].round(decimals=1)

# REPLACE SOME VALUES
df_HowDidYouHearAboutUs_count['_ed5be3a0_How_did_you_hear_about_us_'] = df_HowDidYouHearAboutUs_count['_ed5be3a0_How_did_you_hear_about_us_'].replace(['Other: please specify'],'Other')


# COUNT MEMBERSHIP (FIELD Most Recent Membership Subscription:Type name)

# CLEAN EXPIRED MEMBERSHIP
selected_columns = df[['ID', 'Most Recent Membership Subscription:Expires at', 'Most Recent Membership Subscription:Type name']]
df_MembershipCleaned = selected_columns.copy()
df_MembershipCleaned['Most Recent Membership Subscription:Expires at'] = df_MembershipCleaned['Most Recent Membership Subscription:Expires at'].astype(str)

df_MembershipCleaned['Most Recent Membership Subscription:Expires at'] = df_MembershipCleaned['Most Recent Membership Subscription:Expires at'].str[:-13]

df_MembershipCleaned.loc[df_MembershipCleaned['Most Recent Membership Subscription:Expires at'] < str(today), 'Most Recent Membership Subscription:Type name'] = np.NaN

df_Membership_count = pd.DataFrame(df_MembershipCleaned.groupby(['Most Recent Membership Subscription:Type name'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Membership_count = df_Membership_count.fillna('Basic Membership')

df_Membership_count['Percent'] = (df_Membership_count['Total'] / df_Membership_count['Total'].sum()) * 100
df_Membership_count['Percent'] = df_Membership_count['Percent'].round(decimals=1)


# COUNT EXPERIENCE (How many years have you been in practice)
df_Experience_count = pd.DataFrame(df.groupby(['_83fb023d_How_many_years_have_you_been_in_practice_'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Experience_count = df_Experience_count.fillna('Unknow')

df_Experience_count['Percent'] = (df_Experience_count['Total'] / df_Experience_count['Total'].sum()) * 100
df_Experience_count['Percent'] = df_Experience_count['Percent'].round(decimals=1)


# COUNT DEGREES (CUSTOM FIELD _a7634d0d_Professional_Designation__Degree__)
df_tempDegrees = pd.DataFrame(pd.melt(df['_a7634d0d_Professional_Designation__Degree__'].str.split(', ', expand=True))['value'])

df_Degrees_count = pd.DataFrame(df_tempDegrees.groupby(['value'], dropna=False).size(), columns=['Total'])\
    .reset_index()
df_Degrees_count = df_Degrees_count.fillna('AZERTY')

df_Degrees_count['Percent'] = (df_Degrees_count['Total'] / df.shape[0]) * 100
df_Degrees_count['Percent'] = df_Degrees_count['Percent'].round(decimals=2)

# EMPTY VALUES
degreesEmpty = df['_a7634d0d_Professional_Designation__Degree__'].isna().sum()
degreesEmptyPercent = round((degreesEmpty / df.shape[0]) * 100, 2)

# REPLACE EMPTY VALUES AND SORT
df_Degrees_count.loc[(df_Degrees_count['value'] == 'AZERTY')] = [['Unknow', degreesEmpty, degreesEmptyPercent]]
df_Degrees_count = df_Degrees_count.sort_values(['Total'], ascending=False)


# EXCEL FILE
writer = pd.ExcelWriter(outputExcelFile, engine='xlsxwriter')

df_ActivationCount.to_excel(writer, index=False, sheet_name='Registrations', header=True)
df_Country_count.to_excel(writer, index=False, sheet_name='Countries', header=['Country', 'Total', '%'])
df_AreasCount.to_excel(writer, index=False, sheet_name='Areas', header=['Area', 'Total', '%'])
df_Categories_count.to_excel(writer, index=False, sheet_name='Categories', header=['Category', 'Total', '%'])
df_Specialties_count.to_excel(writer, index=False, sheet_name='Specialties', header=['Specialty', 'Total', '%'])
df_SpecialtiesPerCountry_count.to_excel(writer, index=False, sheet_name='Specialties per country', header=['Country', 'Specialty', 'Total', '%'])
df_Industries_count.to_excel(writer, index=False, sheet_name='Expertise & Interests', header=['Expertise or Interest', 'Total', '%'])
df_Groups_count.to_excel(writer, index=False, sheet_name='Groups', header=['Group', 'Total', '%'])
df_Email_DNS_count.to_excel(writer, index=False, sheet_name='Email domains', header=['Email domain', 'Total', '%'])
df_HowDidYouHearAboutUs_count.to_excel(writer, index=False, sheet_name='How Did You Hear', header=['How did you hear about us (known)', 'Total', '%'])
df_Membership_count.to_excel(writer, index=False, sheet_name='Current memberships', header=['Current memberships', 'Total', '%'])
df_Experience_count.to_excel(writer, index=False, sheet_name='Experience', header=['Experience', 'Total', '%'])
df_Degrees_count.to_excel(writer, index=False, sheet_name='Degrees', header=['Degree', 'Total', '%'])

writer.save()

# EXCEL FILTERS
workbook = openpyxl.load_workbook(outputExcelFile)
sheetsLits = workbook.sheetnames

for sheet in sheetsLits:
    if sheet == 'Registrations':
        continue
    worksheet = workbook[sheet]
    FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    workbook.save(outputExcelFile)

# EXCEL COLORS
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
        workbook.save(outputExcelFile)

# EXCEL COLUMN SIZE
for sheet in sheetsLits:
    for cell in workbook[sheet][1]:
        if get_column_letter(cell.column) == 'A':
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 30
        else:
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 10
        workbook.save(outputExcelFile)

workbook['How Did You Hear'].column_dimensions['A'].width = 40
workbook['Current memberships'].column_dimensions['A'].width = 45
workbook.save(outputExcelFile)

# EXCEL FREEZE TOP ROW
for sheet in sheetsLits:
    if sheet == 'Registrations':
        continue
    if sheet == 'Areas':
        continue
    if sheet == 'Categories':
        continue
    if sheet == 'Groups':
        continue
    if sheet == 'How Did You Hear':
        continue
    if sheet == 'Current memberships':
        continue
    if sheet == 'Experience':
        continue
    if sheet == 'Degrees':
        continue
    worksheet = workbook[sheet]
    worksheet.freeze_panes = 'A2'

# CHART USERS STATUS
activationLabel.pop()
activationValue = []
activationValue.extend([activeUsers, nonActiveUsers, blockedConfirmed, blockedUnconfirmed])
chartLegendPercent = df_ActivationCount['%'].tolist()

explodeValues = []
for i in activationValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(activationLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig1 = plt.figure()
plt.pie(activationValue, labels=None, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('User status', pad=20, fontsize=15)
plt.legend(legendLabels, loc='best', fontsize=10)

fig1.savefig(workDirectory+'myplot1.png', dpi=70)
plt.clf()

im = Image.open(workDirectory+'myplot1.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot1.png')

# INSERT CHART IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot1.png')
img.anchor = 'A8'

workbook['Registrations'].add_image(img)
workbook.save(outputExcelFile)


# CHART CATEGORIES
chartLabel = df_Categories_count['Categories'].tolist()
chartLegendLabel = df_Categories_count['Categories'].tolist()
chartValue = df_Categories_count['Total'].tolist()
chartLegendPercent = df_Categories_count['Percent'].tolist()

chartLabel[-1] = ''

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig2 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Categories', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig2.savefig(workDirectory+'myplot2.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot2.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot2.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot2.png')
img.anchor = 'E4'

workbook['Categories'].add_image(img)
workbook.save(outputExcelFile)


# CHART HOW DID YOU HEAR ABOUT US (CUSTOM FIELD How_did_you_hear_about_us_)
chartLabel = df_HowDidYouHearAboutUs_count['_ed5be3a0_How_did_you_hear_about_us_'].tolist()
chartLegendLabel = df_HowDidYouHearAboutUs_count['_ed5be3a0_How_did_you_hear_about_us_'].tolist()
chartValue = df_HowDidYouHearAboutUs_count['Total'].tolist()
chartLegendPercent = df_HowDidYouHearAboutUs_count['Percent'].tolist()

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig3 = plt.figure()
plt.pie(chartValue, labels=None, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)

plt.axis('equal')
plt.title('How did you hear about us (known)', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig3.savefig(workDirectory+'myplot3.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot3.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot3.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot3.png')
img.anchor = 'E4'

workbook['How Did You Hear'].add_image(img)
workbook.save(outputExcelFile)


# CHART MEMBERSHIP (FIELD Most Recent Membership Subscription:Type name)
chartLabel = df_Membership_count['Most Recent Membership Subscription:Type name'].tolist()
chartValue = df_Membership_count['Total'].tolist()
chartLegendPercent = df_Membership_count['Percent'].tolist()

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig4 = plt.figure()
plt.pie(chartValue, labels=None, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)

plt.axis('equal')
plt.title('Memberships', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig4.savefig(workDirectory+'myplot4.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot4.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot4.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot4.png')
img.anchor = 'E4'

workbook['Current memberships'].add_image(img)
workbook.save(outputExcelFile)


# CHART COUNT REGISTRATIONS BY DATE (FIELD Created at)
chartLabel = df_Created_count['Created'].tolist()
chartValue = df_Created_count['Total'].tolist()

fig5 = plt.figure(figsize=(13,6))
bar_plot = plt.bar(chartLabel, chartValue)

plt.xticks(rotation=30, ha='right')

# HIDE BORDERS
plt.gca().spines['left'].set_color('none')
plt.gca().spines['right'].set_color('none')
plt.gca().spines['top'].set_color('none')

# HIDE TICKS
plt.tick_params(axis='y', labelsize=0, length=0)
plt.yticks([])

# ADD VALUE ON THE TOP OF VERTICAL BARS
def autolabel(rects):
    for idx, rect in enumerate(bar_plot):
        height = rect.get_height()
        plt.text(rect.get_x() + rect.get_width()/2, height,
                chartValue[idx],
                ha='center', va='bottom', rotation=0)

autolabel(bar_plot)

fig5.savefig(workDirectory+'myplot5.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot5.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot5.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot5.png')
img.anchor = 'F2'

workbook['Registrations'].add_image(img)
workbook.save(outputExcelFile)


# CHART GROUPS (FIELD Groups Member:Group Member)
chartLabel = df_Groups_count['value'].tolist()
chartValue = df_Groups_count['Total'].tolist()
chartLegendPercent = df_Groups_count['Percent'].tolist()

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig6 = plt.figure()
plt.pie(chartValue, labels=None, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)

plt.axis('equal')
plt.title('Groups', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig6.savefig(workDirectory+'myplot6.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot6.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot6.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot6.png')
img.anchor = 'F2'

workbook['Groups'].add_image(img)
workbook.save(outputExcelFile)


# MAP COUNTRIES
df_Country_count.set_index('Live Location:Country', inplace=True)

my_values = df_Country_count['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_Country_count['Percent'] = np.digitize(my_values, my_range) - 1

map1 = plt.figure(figsize=(14, 8))

ax = map1.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_countries, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['COUNTRY_HB']
    if shp_ctry not in df_Country_count.index:
        color = '#dddddd'
    else:
        color = scheme[df_Country_count.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map1.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# cb.ax.set_xticklabels([str(round(i, 1)) for i in my_range])
# cb.ax.tick_params(labelsize=7)
# cb.set_label('Percentage', rotation=0)
cb.remove()

map1.savefig(workDirectory+'mymap1.png', dpi=110, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap1.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap1.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap1.png')
img.anchor = 'E2'

workbook['Countries'].add_image(img)
workbook.save(outputExcelFile)


# CHART AREAS
chartLabel = df_AreasCount['continent_stat'].tolist()
chartLegendLabel = df_AreasCount['continent_stat'].tolist()
chartValue = df_AreasCount['Total'].tolist()
chartLegendPercent = df_AreasCount['Percent'].tolist()

chartLabel[-1] = ''

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig7 = plt.figure()
plt.pie(chartValue, labels=None, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Areas', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig7.savefig(workDirectory+'myplot7.png', dpi=90)
plt.clf()

im = Image.open(workDirectory+'myplot7.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot7.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot7.png')
img.anchor = 'A13'

workbook['Areas'].add_image(img)
workbook.save(outputExcelFile)


# MAP AREAS
df_AreasCount.set_index('continent_stat', inplace=True)

my_values = df_AreasCount['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_AreasCount['Percent'] = np.digitize(my_values, my_range) - 1

map3 = plt.figure(figsize=(14, 8))

ax = map3.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_areas, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['continent']
    if shp_ctry not in df_AreasCount.index:
        color = '#dddddd'
    else:
        color = scheme[df_AreasCount.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map3.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

cb.remove()

map3.savefig(workDirectory+'mymap3.png', dpi=90, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap3.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap3.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap3.png')
img.anchor = 'H2'

workbook['Areas'].add_image(img)
workbook.save(outputExcelFile)


# CHART EXPERIENCE (CUSTOM FIELD How many years have you been in practice)
chartLabel = df_Experience_count['_83fb023d_How_many_years_have_you_been_in_practice_'].tolist()
chartLegendLabel = df_Experience_count['_83fb023d_How_many_years_have_you_been_in_practice_'].tolist()
chartValue = df_Experience_count['Total'].tolist()
chartLegendPercent = df_Experience_count['Percent'].tolist()

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig8 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)

plt.axis('equal')
plt.title('Experiences', pad=20, fontsize=15)

plt.legend(legendLabels, loc='lower right', fontsize=8)

fig8.savefig(workDirectory+'myplot8.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot8.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot8.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot8.png')
img.anchor = 'E4'

workbook['Experience'].add_image(img)
workbook.save(outputExcelFile)


# CHART DEGREES (CUSTOM FIELD _a7634d0d_Professional_Designation__Degree__)
chartLabel = df_Degrees_count['value'].tolist()
chartLegendLabel = df_Degrees_count['value'].tolist()
chartValue = df_Degrees_count['Total'].tolist()
chartLegendPercent = df_Degrees_count['Percent'].tolist()

explodeValues = []
for i in chartValue:
    explodeValues.append(0.05)

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig10 = plt.figure()
plt.pie(chartValue, labels=None, colors=colors, explode=explodeValues, autopct=None, shadow=False, startangle=90)

plt.axis('equal')
plt.title('Degrees', pad=20, fontsize=15)

plt.legend(legendLabels, loc='lower right', fontsize=8)

fig10.savefig(workDirectory+'myplot10.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot10.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot10.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot10.png')
img.anchor = 'E4'

workbook['Degrees'].add_image(img)
workbook.save(outputExcelFile)


# REMOVE PICTURES
os.remove(workDirectory+'myplot1.png')
os.remove(workDirectory+'myplot2.png')
os.remove(workDirectory+'myplot3.png')
os.remove(workDirectory+'myplot4.png')
os.remove(workDirectory+'myplot5.png')
os.remove(workDirectory+'myplot6.png')
os.remove(workDirectory+'myplot7.png')
os.remove(workDirectory+'myplot8.png')
os.remove(workDirectory+'myplot10.png')
os.remove(workDirectory+'mymap1.png')
os.remove(workDirectory+'mymap3.png')


# TERMINAL OUTPUTS AND TESTS
print(tab(df_Degrees_count, headers='keys', tablefmt='psql', showindex=False))
print(today)
print("OK, export done!")